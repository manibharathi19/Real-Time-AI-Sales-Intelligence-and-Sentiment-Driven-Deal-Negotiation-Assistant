# utils.py
import openpyxl
import os

def write_analysis_to_excel(file_name, timestamp, transcription, sentiment, intent, action):
    """
    Writes the transcription analysis to an Excel file.
    If the file already exists, data is appended. Otherwise, a new file is created.
    """
    try:
        if os.path.exists(file_name):
            wb = openpyxl.load_workbook(file_name)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Transcriptions"
            ws.append(["Timestamp", "Transcription", "Sentiment", "Intent", "Action"])

        # Append the data
        ws.append([timestamp, transcription, sentiment, intent, action])
        wb.save(file_name)

    except Exception as e:
        print(f"Error saving Excel file: {e}")

def write_summary_to_excel(summary_file_name, username, summary):
    """
    Writes the post-call summary to an Excel file.
    If the file already exists, data is appended. Otherwise, a new file is created.
    """
    try:
        if os.path.exists(summary_file_name):
            wb = openpyxl.load_workbook(summary_file_name)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Post-call Summaries"
            ws.append(["User Name", "Summary"])

        # Append the data
        ws.append([username, summary])
        wb.save(summary_file_name)

    except Exception as e:
        print(f"Error saving Excel file: {e}")
